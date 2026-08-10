"""SQLite-backed reading store for the Setpoint hub.

Replaces the previous append-only CSV file as the system of record.  The CSV
approach rewrote the whole file to add columns and was read in full by the
dashboard every few seconds, which caused blank-dashboard / corruption issues
under concurrent access and did not scale to long-term logging.

Design notes
------------
* WAL journal mode lets the dashboard read while probes are still writing,
  without readers ever seeing a half-written file.
* One connection per thread (``threading.local``) so Flask/waitress worker
  threads never share a connection object.  A single write lock serialises
  writers to avoid "database is locked" errors under load.
* Every row stores both the human-readable local ISO ``ts`` and an integer
  ``epoch`` so time-window queries are index-backed and fast regardless of how
  much history has accumulated.
"""
from __future__ import annotations

import csv as _csv
import datetime
import logging
import sqlite3
import threading
import time
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from core.units import c_to_f

log = logging.getLogger("hub.db")

# Columns exposed to the rest of the app.  ``timestamp`` is aliased from the
# ``ts`` column so existing dashboard code keeps working unchanged.
_SELECT_COLS = "ts AS timestamp, temperature_c, temperature_f, probe_id"


def _reading_filters(cutoff=None, probe_id=None, site=None):
    """Conjunction + params for the readings filters the dashboard queries share.

    Returns ``(clauses, params)`` rather than a finished ``WHERE`` string because
    one caller needs the bare conjunction: ``latest_per_probe`` appends it after
    its own ``probe_id = ?`` in the per-probe seek. :func:`_where` renders the
    ordinary case.

    The two easy-to-get-wrong parts, in one place instead of five:

    * ``site is not None``, not truthiness — ``''`` is a real site (this hub's
      own probes) and must filter rather than mean "no filter". Getting that
      inconsistent across siblings is how the chart ends up showing six stores
      while the cards beside it correctly show one.
    * ``params`` order must track ``clauses`` order, which is why they are built
      together and returned together.
    """
    clauses, params = [], []
    if cutoff is not None:
        clauses.append("epoch >= ?"); params.append(cutoff)
    if probe_id:
        clauses.append("probe_id = ?"); params.append(probe_id)
    if site is not None:
        clauses.append("site = ?"); params.append(site)
    return clauses, tuple(params)


def _where(clauses) -> str:
    return ("WHERE " + " AND ".join(clauses)) if clauses else ""


_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value) -> str:
    """Neutralise spreadsheet formula injection in an exported CSV cell.

    A cell that begins with ``= + - @`` (or a leading tab/CR) is treated as a
    formula by Excel/Sheets; prefixing it with a single quote makes it plain
    text. Applied to the free-form string columns of the export.
    """
    s = "" if value is None else str(value)
    if s and s[0] in _FORMULA_LEAD:
        return "'" + s
    return s


def _xlsx_safe(value) -> str:
    """Neutralise formula injection for a string written into an .xlsx cell.

    openpyxl treats a leading ``=`` as a formula exactly as Excel does for CSV,
    so the guard is identical to :func:`_csv_safe`; kept as a named alias for
    clarity at the xlsx call sites.
    """
    return _csv_safe(value)


class ExportTooLargeForXlsx(Exception):
    """Raised when a requested .xlsx export would exceed Excel's row limit.

    Carries the actual and maximum row counts so the caller can tell the user to
    narrow the date range (or use the CSV export, which has no such limit).
    """

    def __init__(self, rows: int, max_rows: int):
        self.rows = rows
        self.max_rows = max_rows
        super().__init__(
            f"{rows:,} rows exceeds Excel's limit of {max_rows:,}; "
            f"narrow the date range or use a CSV export.")


def iso_to_epoch(ts: str) -> float:
    """Convert a local-naive ISO timestamp to a POSIX epoch (fractional seconds).

    Naive timestamps are interpreted as local machine time, matching how the
    rest of the app stores them.  Returns the current epoch if parsing fails so
    a malformed timestamp never blocks an insert.

    The value is a float so sub-second timestamps (high-rate logging) keep their
    precision. SQLite's flexible typing stores a whole-second epoch as an INTEGER
    (lossless) and a fractional one as REAL in the same ``epoch`` column, so this
    is backward-compatible with existing integer rows and the epoch index.
    """
    try:
        s = str(ts).strip().rstrip("Z")
        return datetime.datetime.fromisoformat(s).timestamp()
    except Exception:
        return time.time()


def utc_iso(epoch) -> str:
    """Format an epoch as an unambiguous ISO-8601 UTC string.

    Carries millisecond precision when the value has it (sub-second/high-rate
    logging), else keeps the clean seconds format.

    Module level, not a private method, because it is the inverse of
    :func:`core.storage.absolute_epoch` and both sides of the hub-to-hub wire
    need it: the exports render a UTC column with it, and the upstream forwarder
    stamps its payload with it so the receiving hub can recover the exact
    instant instead of re-deriving one from a local-naive string in ITS OWN
    timezone. ``Database._utc_string`` remains as the in-class alias.
    """
    utc_dt = datetime.datetime.fromtimestamp(float(epoch), tz=datetime.timezone.utc)
    if utc_dt.microsecond:
        return utc_dt.strftime("%Y-%m-%dT%H:%M:%S.") + f"{utc_dt.microsecond // 1000:03d}Z"
    return utc_dt.strftime("%Y-%m-%dT%H:%M:%SZ")


# A wall clock earlier than this is treated as not-yet-synced and is NOT trusted
# for the destructive startup future-reading purge (see delete_future_readings).
# Imported from core.storage, which uses the same floor to decide whether to trust
# the hub's clock over a probe's timestamp on ingest — one definition so the two
# guards can never disagree about when the clock became believable.
from core.storage import CLOCK_TRUSTWORTHY_FLOOR_EPOCH as _CLOCK_TRUSTWORTHY_FLOOR_EPOCH
# The events log accepts a timestamp straight off the wire (a forwarding hub
# sends the instant its OWN alert engine decided something). Both helpers are
# needed to store it the way every other row is stored: one hub-local wall clock
# in ``ts``, and the true instant in ``epoch``.
from core.storage import _to_local_naive as _storage_to_local_naive
from core.storage import absolute_epoch as _storage_absolute_epoch


class Database:
    def __init__(self, path: str | Path):
        self.path = str(path)
        self._write_lock = threading.Lock()
        self._local = threading.local()
        self._init_schema()

    # -- connection management -------------------------------------------------
    def _conn(self) -> sqlite3.Connection:
        conn = getattr(self._local, "conn", None)
        if conn is None:
            conn = sqlite3.connect(self.path, check_same_thread=False, timeout=30.0)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.execute("PRAGMA busy_timeout=30000")
            self._local.conn = conn
        return conn

    def _init_schema(self) -> None:
        conn = self._conn()
        with self._write_lock:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS readings (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    ts            TEXT    NOT NULL,
                    epoch         INTEGER NOT NULL,
                    temperature_c REAL    NOT NULL,
                    temperature_f REAL    NOT NULL,
                    probe_id      TEXT    NOT NULL DEFAULT '',
                    humidity_pct  REAL,
                    vpd_kpa       REAL,
                    battery_pct   REAL
                )
                """
            )
            # Forward-migrate a pre-humidity/pre-battery database: ADD COLUMN is a
            # no-op error if the column already exists, so guard each one
            # independently.
            for col in ("humidity_pct", "vpd_kpa", "battery_pct"):
                try:
                    conn.execute(f"ALTER TABLE readings ADD COLUMN {col} REAL")
                except sqlite3.OperationalError:
                    pass  # column already present
            # ``site`` names which building a reading came from. Empty for every
            # locally-ingested reading (the overwhelmingly common case, and what
            # a single-site hub stays forever); set only on rows an upstream
            # forwarder pushed here from another hub. TEXT, not REAL, so it gets
            # its own ALTER — see core/forwarder.py and docs/MULTI_SITE.md.
            try:
                conn.execute("ALTER TABLE readings ADD COLUMN site TEXT NOT NULL DEFAULT \'\'")
            except sqlite3.OperationalError:
                pass
            conn.execute("CREATE INDEX IF NOT EXISTS idx_readings_epoch ON readings(epoch)")
            # PARTIAL index, and the `WHERE site != ''` is the whole point: on a
            # single-site hub no row qualifies, so the index is empty and costs
            # nothing in disk or insert time — measured identical DB size at 1.5M
            # rows. On an HQ hub it indexes only the forwarded rows and turns
            # ``sites()`` from a full table scan into a handful of seeks.
            conn.execute("CREATE INDEX IF NOT EXISTS idx_readings_site "
                         "ON readings(site) WHERE site != ''")
            # Small durable key/value scratchpad. The upstream forwarder keeps
            # its high-water mark here so a hub restart neither re-sends the
            # whole history nor silently skips the rows written while it was
            # down. Deliberately generic — a second consumer beats a second table.
            conn.execute("CREATE TABLE IF NOT EXISTS meta (k TEXT PRIMARY KEY, v TEXT NOT NULL)")
            self._ensure_reading_uniqueness(conn)
            # Alert-lifecycle event log (threshold breach/recovery, probe
            # offline/online, rate-of-change) — powers the dashboard's recent
            # events feed without re-deriving history from raw readings.
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS events (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    ts            TEXT    NOT NULL,
                    epoch         INTEGER NOT NULL,
                    kind          TEXT    NOT NULL,
                    probe_id      TEXT    NOT NULL DEFAULT '',
                    temperature_c REAL,
                    limit_c       REAL
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_events_epoch ON events(epoch)")
            # Which hub RECORDED this event. '' is this one; a store's label on a
            # row its hub forwarded here. Readings answer "what was the
            # temperature"; events answer "what did that store's own hub decide
            # about it, against its own thresholds" — a different record, and the
            # one a food-safety auditor actually asks for.
            try:
                conn.execute("ALTER TABLE events ADD COLUMN site TEXT NOT NULL DEFAULT ''")
            except sqlite3.OperationalError:
                pass
            # Idempotency for FORWARDED events only — hence `WHERE site != ''`,
            # and that restriction is load-bearing rather than an optimisation.
            # Event epochs are whole seconds (readings carry milliseconds), so a
            # blanket unique key would treat two genuinely distinct same-second
            # events for one probe as a replay and silently drop the second. The
            # only place a replay can actually happen is a forwarded batch re-sent
            # after a dropped response, so constrain exactly that and leave every
            # locally-recorded event alone. Being partial also makes the index
            # empty on a hub nobody forwards to, which is nearly all of them.
            try:
                conn.execute(
                    "CREATE UNIQUE INDEX IF NOT EXISTS idx_events_fwd_dedup "
                    "ON events(kind, probe_id, epoch, site) WHERE site != ''")
            except sqlite3.IntegrityError:
                # A log that already holds forwarded duplicates: keep the history,
                # lose only the constraint.
                conn.execute(
                    "CREATE INDEX IF NOT EXISTS idx_events_fwd_dedup "
                    "ON events(kind, probe_id, epoch, site) WHERE site != ''")
            conn.commit()

    # Set in ``meta`` once a database is proven to hold genuinely-distinct
    # readings that share a (probe_id, epoch, site), so uniqueness can never be
    # enforced on it. Without the marker every startup re-ran the dedupe scan
    # below — a full-table GROUP BY plus a failed index build — on precisely the
    # oldest and largest databases.
    _NO_UNIQUE_KEY = "schema.readings_not_unique"
    _UNIQUE_INDEX = "idx_readings_probe_epoch_site_uniq"

    def _ensure_reading_uniqueness(self, conn: sqlite3.Connection) -> None:
        """Enforce one physical reading per (probe, instant, site).

        The UNIQUE index makes ingest idempotent: a re-sent bulk chunk (a dropped
        ACK on an ``/ingest_csv`` flush) or a replayed single POST cannot create
        duplicate rows that would inflate COUNT/AVG/min-max and the CSV exports.
        Every reading carries a millisecond-precision timestamp — the probe's
        ``nowIso()``, or the hub's own ms stamp when a payload omits one — so two
        DISTINCT readings never share an epoch; only a byte-identical replay
        collides, and ``INSERT OR IGNORE`` drops it.

        ``site`` is in the key, and its POSITION is the whole trick. Two stores
        whose operators both named a probe "walkin" produce the same
        (probe_id, epoch) at head office, and a two-column key made INSERT OR
        IGNORE silently discard the second store's reading — a hole in a
        food-safety record, created by nothing worse than two people picking the
        same obvious name. Adding site as the THIRD column fixes that while
        leaving (probe_id, epoch) as the leading prefix, so the per-probe "newest
        row" seek in :meth:`latest_per_probe` still uses this index as before.

        Caller holds the write lock.
        """
        have_unique = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='index' AND name=?",
            (self._UNIQUE_INDEX,)).fetchone() is not None
        if have_unique:
            return  # already migrated — the overwhelmingly common path
        gave_up = conn.execute("SELECT 1 FROM meta WHERE k=?",
                               (self._NO_UNIQUE_KEY,)).fetchone() is not None
        if gave_up:
            # A previous run already proved this database can't take the unique
            # index. Just make sure the non-unique one serving the same query
            # patterns is present, and skip the scan.
            conn.execute("CREATE INDEX IF NOT EXISTS idx_readings_probe_epoch "
                         "ON readings(probe_id, epoch)")
            return

        conn.execute("DROP INDEX IF EXISTS idx_readings_probe_epoch")
        conn.execute("DROP INDEX IF EXISTS idx_readings_probe_epoch_uniq")
        try:
            conn.execute(f"CREATE UNIQUE INDEX {self._UNIQUE_INDEX} "
                         "ON readings(probe_id, epoch, site)")
            return
        except sqlite3.IntegrityError:
            pass

        # A pre-existing database from the older non-idempotent path may hold
        # rows that share a (probe_id, epoch). Collapse ONLY byte-identical
        # duplicates — every reading column equal — keeping the earliest id. Rows
        # that share an epoch but differ in any value are DISTINCT measurements
        # (old whole-second stamping of a sub-1 s cadence, before ms timestamps)
        # and MUST be preserved: grouping on the full row, not just
        # (probe_id, epoch), is what keeps this from deleting real readings.
        conn.execute(
            "DELETE FROM readings WHERE id NOT IN ("
            " SELECT MIN(id) FROM readings GROUP BY probe_id, epoch, site, "
            " temperature_c, temperature_f, humidity_pct, vpd_kpa, battery_pct)"
        )
        try:
            conn.execute(f"CREATE UNIQUE INDEX {self._UNIQUE_INDEX} "
                         "ON readings(probe_id, epoch, site)")
        except sqlite3.IntegrityError:
            # Genuinely-distinct readings still collide, so uniqueness cannot be
            # enforced without discarding real data. Keep the old NON-unique
            # index instead: ingest idempotency just doesn't apply retroactively
            # to this legacy DB (new readings carry ms-precision epochs and don't
            # collide), which is strictly better than losing history. Record the
            # verdict so the next startup doesn't repeat this scan.
            conn.execute("CREATE INDEX IF NOT EXISTS idx_readings_probe_epoch "
                         "ON readings(probe_id, epoch)")
            conn.execute("INSERT OR REPLACE INTO meta (k, v) VALUES (?, '1')",
                         (self._NO_UNIQUE_KEY,))
            log.warning(
                "This database holds readings that share a probe id and instant "
                "but differ in value (pre-millisecond timestamps), so duplicate "
                "ingests cannot be de-duplicated retroactively. New readings are "
                "unaffected.")

    # -- writes ----------------------------------------------------------------
    def append(self, ts: str, t_c: float, t_f: float, probe_id: str = "",
               humidity: float | None = None, vpd: float | None = None,
               battery: float | None = None, epoch: float | None = None,
               site: str = "") -> None:
        # ``epoch`` lets the caller supply the TRUE instant when the incoming
        # timestamp carried timezone info. Re-deriving it from the local-naive
        # ``ts`` is ambiguous during the DST fall-back hour, where two readings an
        # hour apart share one wall time — they would collide on
        # UNIQUE(probe_id, epoch, site) and sort interleaved. See storage.absolute_epoch.
        epoch = iso_to_epoch(ts) if epoch is None else float(epoch)
        conn = self._conn()
        with self._write_lock:
            conn.execute(
                "INSERT OR IGNORE INTO readings (ts, epoch, temperature_c, temperature_f, probe_id, "
                "humidity_pct, vpd_kpa, battery_pct, site) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (str(ts), epoch, float(t_c), float(t_f), probe_id or "",
                 (float(humidity) if humidity is not None else None),
                 (float(vpd) if vpd is not None else None),
                 (float(battery) if battery is not None else None),
                 site or ""),
            )
            conn.commit()

    def record_event(self, kind: str, probe_id: str, temperature_c=None,
                     limit=None, ts=None, site: str = "") -> None:
        """Append one alert-lifecycle event to the events log.

        ``kind`` is one of ``'high' 'low' 'recovery' 'offline' 'online' 'rate'``;
        ``ts`` defaults to the current local time.  Best-effort by design: the
        alert cycle that emits an event must never be broken by an unrecordable
        one, so bad numeric fields are coerced to NULL, a blank kind is skipped,
        and any storage error is swallowed rather than raised to the caller.

        A ``ts`` carrying explicit timezone info (a ``Z`` or a ``+HH:MM``
        offset) is handled the way an ingested reading is: converted to this
        hub's wall clock for the ``ts`` column, with the TRUE instant kept for
        ``epoch``. That path is the multi-site forwarder — a store hub sends the
        moment its own alert engine decided something, and head office in
        another timezone must not re-read that wall clock as its own.
        """
        def _f(v):
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None
        try:
            kind = str(kind or "").strip()
            if not kind:
                return  # nothing meaningful to record
            raw_ts = str(ts) if ts else ""
            if raw_ts:
                # None for a naive stamp — then iso_to_epoch's local reading of
                # it is the best that is knowable, exactly as for readings.
                exact = _storage_absolute_epoch(raw_ts)
                ts = _storage_to_local_naive(raw_ts)
            else:
                exact = None
                ts = datetime.datetime.now().isoformat(timespec="seconds")
            epoch = int(iso_to_epoch(ts) if exact is None else float(exact))
            conn = self._conn()
            with self._write_lock:
                conn.execute(
                    "INSERT OR IGNORE INTO events "
                    "(ts, epoch, kind, probe_id, temperature_c, limit_c, site) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (ts, epoch, kind, str(probe_id or ""),
                     _f(temperature_c), _f(limit), site or ""),
                )
                conn.commit()
        except Exception:
            pass  # an event is telemetry — never worth failing the caller over

    def bulk_insert(self, rows, site: str = "") -> int:
        """Insert many readings in one transaction.

        Accepts either ``(ts, t_c, t_f, probe_id)`` or the richer
        ``(ts, t_c, t_f, probe_id, humidity, vpd, battery)`` — the short form is
        kept so the legacy-CSV migration and existing callers are unaffected.
        Carrying the optional telemetry matters for the bulk backlog drain: a
        grow probe (SHT4x) that reconnects after an outage would otherwise have
        every buffered humidity/VPD/battery value silently dropped, even though
        the live path stores them, so an outage would leave a temperature-only
        hole in exactly the data that niche buys the product for.

        Used for the legacy-CSV migration and the probe's bulk backlog drain
        (``/api/ingest_csv``) so importing tens of thousands of rows is a single
        commit rather than one per row.  Uses ``INSERT OR IGNORE`` against the
        UNIQUE(probe_id, epoch, site) index, so a re-sent chunk (a dropped ACK on a
        flush) does not duplicate already-stored readings.  Returns the number of
        rows actually inserted (replayed duplicates are skipped, not counted).
        """
        def _f(v):
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None

        params = []
        for row in rows:
            ts, t_c, t_f, pid = row[0], row[1], row[2], row[3]
            hum, vpd, bat = (list(row[4:7]) + [None, None, None])[:3] if len(row) > 4 \
                else (None, None, None)
            # Optional 8th element: the TRUE epoch, when the source timestamp
            # carried timezone info (see append()/storage.absolute_epoch — the
            # DST fall-back hour is otherwise ambiguous).
            exact = row[7] if len(row) > 7 else None
            epoch = iso_to_epoch(ts) if exact is None else float(exact)
            params.append((str(ts), epoch, float(t_c), float(t_f),
                           (pid or ""), _f(hum), _f(vpd), _f(bat), site or ""))
        if not params:
            return 0
        conn = self._conn()
        with self._write_lock:
            before = conn.total_changes
            conn.executemany(
                "INSERT OR IGNORE INTO readings (ts, epoch, temperature_c, temperature_f, "
                "probe_id, humidity_pct, vpd_kpa, battery_pct, site) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                params,
            )
            conn.commit()
            inserted = conn.total_changes - before
        return inserted

    # -- reads -----------------------------------------------------------------
    def count(self) -> int:
        row = self._conn().execute("SELECT COUNT(*) AS n FROM readings").fetchone()
        return int(row["n"]) if row else 0

    def has_any(self) -> bool:
        """True if the store holds at least one reading. ``EXISTS`` stops at the
        first row, so this is O(1) — unlike ``count()`` (a full scan) it is cheap
        to call on every dashboard tick just to test emptiness."""
        row = self._conn().execute("SELECT EXISTS(SELECT 1 FROM readings) AS e").fetchone()
        return bool(row and row["e"])

    def has_probe_prefix(self, prefix: str) -> bool:
        """True if any reading's ``probe_id`` starts with ``prefix``.

        The GLOB prefix pattern rewrites to a range seek on whichever index leads
        with ``probe_id`` — ``idx_readings_probe_epoch_site_uniq``, or the legacy
        ``idx_readings_probe_epoch`` on a database that could not take the unique
        one (verified via EXPLAIN QUERY PLAN in the tests) — so answering "is demo
        data loaded?" on every settings render is O(log N) instead of
        scanning/grouping the whole readings table the way
        ``last_reading_epoch_per_probe`` does.  ``prefix`` is expected to be a
        plain probe-id token (no GLOB metacharacters).
        """
        if not prefix:
            return self.has_any()
        row = self._conn().execute(
            "SELECT EXISTS(SELECT 1 FROM readings WHERE probe_id GLOB ?) AS e",
            (str(prefix) + "*",),
        ).fetchone()
        return bool(row and row["e"])

    def latest(self) -> Optional[dict]:
        """Most recent reading overall, or None if the table is empty."""
        row = self._conn().execute(
            f"SELECT {_SELECT_COLS} FROM readings ORDER BY epoch DESC, id DESC LIMIT 1"
        ).fetchone()
        return dict(row) if row else None

    def last_reading_epoch_per_probe(self, window_seconds: Optional[int] = None,
                                     site: Optional[str] = None) -> dict:
        """Map ``probe_id -> epoch of its most recent reading`` within the window.

        Used for offline detection: a probe whose newest reading is older than
        the offline threshold has gone silent.  The window bounds which probes
        are tracked, so long-retired probes drop out instead of alerting forever.

        ``site`` narrows to one forwarding store on an HQ hub; ``None`` (the
        default, and what offline detection passes) covers every site.
        """
        conn = self._conn()
        cutoff = self._cutoff(window_seconds)
        clauses, params = _reading_filters(cutoff, site=site)
        where = _where(clauses)
        rows = conn.execute(
            f"SELECT probe_id, MAX(epoch) AS last_epoch FROM readings {where} GROUP BY probe_id",
            params,
        ).fetchall()
        return {r["probe_id"]: int(r["last_epoch"]) for r in rows if r["probe_id"]}

    def _cutoff(self, window_seconds: Optional[int]) -> Optional[int]:
        if not window_seconds:
            return None
        return int(time.time()) - int(window_seconds)

    def window_df(self, window_seconds: Optional[int] = None, max_points: int = 6000,
                  probe_id: Optional[str] = None,
                  site: Optional[str] = None) -> pd.DataFrame:
        """Return readings within a rolling window as a DataFrame.

        When the window contains more than ``max_points`` rows the result is
        uniformly downsampled in SQL (per-probe row position) so the dashboard
        stays responsive even with millions of historical readings.  Statistics
        are computed separately on the full window via :meth:`window_stats`, so
        downsampling only affects plot density, never the reported min/max/avg.

        When ``probe_id`` is given, BOTH the row scan and the COUNT that sizes the
        downsample stride are restricted to that probe — the dashboard's focus
        mode must derive the stride from the FOCUSED probe's own volume, not the
        global all-probe total, or a quiet probe gets decimated to a point or two
        while a chatty neighbour dominates the count.
        """
        conn = self._conn()
        cutoff = self._cutoff(window_seconds)
        clauses, params = _reading_filters(cutoff, probe_id, site)
        where = _where(clauses)

        total = conn.execute(f"SELECT COUNT(*) AS n FROM readings {where}", params).fetchone()["n"]
        if total == 0:
            return pd.DataFrame(columns=["timestamp", "temperature_c", "temperature_f", "probe_id"])

        stride = max(1, (total + max_points - 1) // max_points)
        if stride > 1:
            # Downsample by per-probe ROW POSITION, not by `id % stride`. Primary
            # keys are not contiguous (multiple writers interleave, delete_probe/
            # purge leave gaps), so `id % stride` biases the sample per probe and
            # can select ZERO rows (blank chart) or drop entire probes. Numbering
            # each probe's rows newest-first and keeping every stride-th one means
            # rn=1 (the live tip) is always kept, every probe is represented, and
            # the result is never empty. Point count stays near max_points (a
            # multi-probe window can exceed it by up to one point per probe —
            # bounded by the probe-registry cap and harmless for the chart).
            sql = (
                "SELECT timestamp, temperature_c, temperature_f, probe_id FROM ("
                "  SELECT ts AS timestamp, epoch, id, temperature_c, temperature_f, probe_id,"
                "         ROW_NUMBER() OVER (PARTITION BY probe_id ORDER BY epoch DESC, id DESC) AS rn"
                f"  FROM readings {where}"
                ") WHERE (rn - 1) % ? = 0 ORDER BY epoch ASC, id ASC"
            )
            rows = conn.execute(sql, params + (stride,)).fetchall()
        else:
            sql = f"SELECT {_SELECT_COLS} FROM readings {where} ORDER BY epoch ASC"
            rows = conn.execute(sql, params).fetchall()

        return pd.DataFrame([{"timestamp": r["timestamp"], "temperature_c": r["temperature_c"],
                              "temperature_f": r["temperature_f"], "probe_id": r["probe_id"]}
                             for r in rows],
                            columns=["timestamp", "temperature_c", "temperature_f", "probe_id"])

    def window_stats(self, window_seconds: Optional[int] = None,
                     probe_id: Optional[str] = None,
                     site: Optional[str] = None) -> dict:
        """Accurate min/max/avg/count over the full (un-downsampled) window.

        When ``probe_id`` is given, the stats cover only that probe — used by the
        dashboard's "focus one probe" mode so the min/max/avg describe the single
        selected probe instead of all probes mixed together.
        """
        conn = self._conn()
        cutoff = self._cutoff(window_seconds)
        clauses, params = _reading_filters(cutoff, probe_id, site)
        where = _where(clauses)

        agg = conn.execute(
            f"SELECT COUNT(*) AS n, MIN(temperature_c) AS mn, MAX(temperature_c) AS mx, "
            f"AVG(temperature_c) AS av FROM readings {where}",
            params,
        ).fetchone()
        if not agg or not agg["n"]:
            return {"count": 0, "min": None, "max": None, "avg": None,
                    "min_ts": None, "max_ts": None}

        # probe_id rides along free: these queries already fetch the exact row
        # holding the extreme. Without it an overview tile reads "MIN 37.0 F"
        # with no way to tell whether that was the freezer or the store room —
        # a true number nobody can act on. `site` rides along for the same
        # reason and one more: the tile's red/blue colouring judges the extreme
        # against a threshold, and a forwarded row must be judged by the hub
        # that owns it, not by this one (core.status.limits_for).
        min_row = conn.execute(
            f"SELECT ts, probe_id, site FROM readings {where} "
            f"ORDER BY temperature_c ASC, epoch ASC LIMIT 1", params
        ).fetchone()
        max_row = conn.execute(
            f"SELECT ts, probe_id, site FROM readings {where} "
            f"ORDER BY temperature_c DESC, epoch ASC LIMIT 1", params
        ).fetchone()
        return {
            "count": int(agg["n"]),
            "min": agg["mn"],
            "max": agg["mx"],
            "avg": agg["av"],
            "min_ts": min_row["ts"] if min_row else None,
            "max_ts": max_row["ts"] if max_row else None,
            "min_probe": (min_row["probe_id"] or "") if min_row else "",
            "max_probe": (max_row["probe_id"] or "") if max_row else "",
            "min_site": (min_row["site"] or "") if min_row else "",
            "max_site": (max_row["site"] or "") if max_row else "",
        }

    def stats_per_probe(self, window_seconds: Optional[int] = None,
                        site: Optional[str] = None) -> dict:
        """Per-probe min/max/avg/count over the full (un-downsampled) window.

        Returns ``{probe_id: {count, min, max, avg}}``. Used for the dashboard's
        per-probe statistics breakdown, where a single global average across
        probes of different ranges (a −18 °C freezer + a 22 °C room) would be
        meaningless. Rows with an empty ``probe_id`` are grouped under ``""``.

        ``site`` narrows to one forwarding store on an HQ hub; ``''`` selects
        this hub's own probes, ``None`` (the default) every site at once.
        """
        conn = self._conn()
        cutoff = self._cutoff(window_seconds)
        clauses, params = _reading_filters(cutoff, site=site)
        where = _where(clauses)
        rows = conn.execute(
            f"SELECT probe_id, COUNT(*) AS n, MIN(temperature_c) AS mn, "
            f"MAX(temperature_c) AS mx, AVG(temperature_c) AS av "
            f"FROM readings {where} GROUP BY probe_id",
            params,
        ).fetchall()
        return {
            (r["probe_id"] or ""): {
                "count": int(r["n"]), "min": r["mn"], "max": r["mx"], "avg": r["av"],
            }
            for r in rows if r["n"]
        }

    def latest_per_probe(self, window_seconds: Optional[int] = None,
                         site: Optional[str] = None) -> pd.DataFrame:
        """Latest reading for each probe within the window (for alerts/display).

        Ties on ``epoch`` (two readings in the same second) are broken by
        insertion ``id`` so "latest" is always the most recently stored row.

        Implemented as per-probe index seeks on ``idx_readings_probe_epoch``:
        the distinct probe ids come straight off the index, then each probe's
        newest row is one backward ``ORDER BY epoch DESC LIMIT 1`` seek (``id``
        is the rowid, so the index order breaks epoch ties for free).  That is
        O(probes x log N) per call, replacing a ROW_NUMBER() window scan that
        touched every row in the window on every dashboard tick.

        ``site`` narrows to one forwarding store on an HQ hub; ``''`` selects
        this hub's own probes, ``None`` (the default) every site at once.

        Alerting passes ``''``. It used to pass ``None``, on the reasoning that "a
        breach in store 3 is still a breach when head office happens to be looking
        at store 1" — true, and head office does learn about it, because the store
        forwards the event. What head office cannot do is *judge* it: thresholds
        and reporting intervals are per hub and are not forwarded, so evaluating a
        store's readings against HQ's own config is confidently wrong. An HQ
        running fridges mailed a LOW alarm for every healthy store freezer, and a
        15-minute-cadence probe was called offline against HQ's 300 s window. See
        ``OWN_SITE`` in alert_monitor.py.
        """
        conn = self._conn()
        cutoff = self._cutoff(window_seconds)
        clauses, params = _reading_filters(cutoff, site=site)
        # The probe list comes from a loose index scan over the WHOLE table, not
        # a windowed SELECT DISTINCT. The seeks below were always the cheap half;
        # the DISTINCT was the expensive one, because deduplicating a
        # low-cardinality column means visiting every row in the window — 64 ms
        # of a 78 ms call on a 2M-row store, on every 5 s dashboard tick, from
        # four callbacks and the alert monitor. Walking the index one probe at a
        # time is O(probes x log N), which is what this method always claimed to
        # be.
        #
        # Dropping the window from THIS query changes nothing: each seek still
        # carries the window, and a probe with no row inside it returns None and
        # is skipped below. It costs one extra seek per long-retired probe.
        pids = self.probe_ids()
        cols = ["timestamp", "temperature_c", "temperature_f", "probe_id",
                "humidity_pct", "vpd_kpa", "battery_pct", "site"]
        seek_clause = ("" if not clauses else " AND " + " AND ".join(clauses))
        out = []
        for pid in pids:
            row = conn.execute(
                f"SELECT ts AS timestamp, temperature_c, temperature_f, probe_id, "
                f"humidity_pct, vpd_kpa, battery_pct, site FROM readings "
                f"WHERE probe_id = ?{seek_clause} ORDER BY epoch DESC, id DESC LIMIT 1",
                (pid,) + params,
            ).fetchone()
            if row is not None:
                out.append({k: row[k] for k in cols})
        return pd.DataFrame(out, columns=cols)

    def fetch_readings(self, window_seconds: Optional[int] = None,
                       probe_id: Optional[str] = None,
                       start_epoch: Optional[int] = None,
                       end_epoch: Optional[int] = None,
                       limit: Optional[int] = None,
                       max_points: int = 6000,
                       site: Optional[str] = None) -> list:
        """Return readings as a list of dict rows for the JSON read API.

        Accepts the same filters as :meth:`export_csv` (a rolling
        ``window_seconds``, a single ``probe_id``, and an absolute
        ``start_epoch``/``end_epoch`` range, all AND-ed). Unlike
        :meth:`window_df` the rows include humidity and VPD. The result is
        bounded to at most ``limit`` (default ``max_points``, hard cap 50 000)
        of the most RECENT matching rows, returned oldest-first, so a
        months-long store can never emit an unbounded payload over the API.
        """
        conn = self._conn()
        # Same filters as every export variant, so they share the builder — this
        # was a byte-identical copy, comment included, and the end_epoch float()
        # subtlety it documents is exactly the kind that gets fixed in one copy.
        where, params_list = self._export_where(window_seconds, probe_id,
                                                start_epoch, end_epoch, site)
        try:
            cap = int(limit) if limit else int(max_points)
        except (TypeError, ValueError):
            cap = int(max_points)
        cap = max(1, min(cap, 50000))
        rows = conn.execute(
            f"SELECT ts AS timestamp, temperature_c, temperature_f, probe_id, "
            f"humidity_pct, vpd_kpa, battery_pct FROM readings {where} "
            f"ORDER BY epoch DESC, id DESC LIMIT ?",
            params_list + (cap,),
        ).fetchall()
        out = [dict(r) for r in rows]
        out.reverse()  # oldest-first for charting/time-series consumers
        return out

    def oldest_temp_c_in_window(self, probe_id, window_seconds) -> Optional[float]:
        """The earliest reading's °C within the rolling window for one probe.

        The rate-of-change alert needs the true *window-old* sample. It must NOT
        use ``fetch_readings(...)[0]``: that caps to the most-recent ``max_points``
        rows FIRST and then reverses, so on a high-cadence probe (e.g. 0.5 s over
        a 60 min window = 7200 rows > 6000 cap) its oldest row is only
        ~cap-samples-old, understating the delta and missing a slow rate alert.
        A direct ``ORDER BY epoch ASC LIMIT 1`` (index-backed) is exact.
        """
        cutoff = self._cutoff(window_seconds)
        conn = self._conn()
        where = "WHERE probe_id = ?" + (" AND epoch >= ?" if cutoff is not None else "")
        params = (probe_id,) if cutoff is None else (probe_id, cutoff)
        row = conn.execute(
            f"SELECT temperature_c FROM readings {where} ORDER BY epoch ASC LIMIT 1",
            params,
        ).fetchone()
        return float(row["temperature_c"]) if row else None

    def readings_after(self, probe_id, after_epoch, not_before=None) -> list:
        """``[(epoch, temperature_c), ...]`` ascending, strictly after ``after_epoch``.

        Deliberately not ``fetch_readings``: that caps to the most-recent
        ``max_points`` rows, which is right for a chart and wrong here. The
        missed-excursion scan is looking for a breach that STARTED early in a
        backfill, and a row cap would drop exactly that end of the window.

        There is no upper bound, and that is the point. Bounding it with
        ``last_reading_epoch_per_probe`` — the obvious choice — silently loses the
        newest reading: that helper returns ``int(MAX(epoch))``, and a stored
        epoch carries milliseconds, so ``epoch <= <truncated>`` excludes the very
        row it was meant to include. A run that had just ended then looked like it
        was still open, got left to the live evaluator, and was never reported by
        either. ``not_before`` is the caller's own floor for a long outage.
        """
        clauses = ["probe_id = ?", "epoch > ?"]
        params = [probe_id, float(after_epoch)]
        if not_before is not None:
            clauses.append("epoch >= ?")
            params.append(float(not_before))
        rows = self._conn().execute(
            "SELECT epoch, temperature_c FROM readings "
            f"WHERE {' AND '.join(clauses)} ORDER BY epoch ASC",
            params,
        ).fetchall()
        return [(r["epoch"], r["temperature_c"]) for r in rows]

    def latest_alarm_state_per_probe(self, window_seconds: Optional[int] = None,
                                     site: Optional[str] = None) -> dict:
        """``{probe_id: 'high'|'low'}`` for probes whose newest lifecycle event
        says they are currently in alarm.

        Exists for the probes this hub may not judge for itself. Head office
        holds every store's readings but none of their thresholds, so it cannot
        derive a verdict — but the store forwards the verdict it reached, and
        that is the thing head office should be showing. Without this, HQ drew a
        store freezer sitting in a live breach as "no alarm set" while the store's
        own screen showed "▲ HIGH": two hubs, one probe, one moment, two answers.

        Only ``high``/``low``/``recovery`` are considered, newest per probe wins,
        and ``recovery`` resolves to absent. ``missed`` is deliberately excluded:
        it records an excursion that has already ENDED, so treating it as a live
        alarm would leave a probe looking broken forever after one bad night.
        """
        rows = self.list_events(limit=2000, window_seconds=window_seconds,
                                kinds=("high", "low", "recovery"), site=site)
        out: dict = {}
        for r in rows:                      # newest first
            pid = r.get("probe_id")
            if not pid or pid in out:
                continue
            out[pid] = r.get("kind")
        return {p: k for p, k in out.items() if k in ("high", "low")}

    def max_reading_id(self) -> int:
        """The highest ``readings.id`` in the store, or 0 when empty."""
        row = self._conn().execute("SELECT MAX(id) FROM readings").fetchone()
        return int(row[0]) if row and row[0] is not None else 0

    def readings_since_id(self, after_id, limit=None, site=None) -> list:
        """``[(id, probe_id, epoch, temperature_c), ...]`` that ARRIVED after ``after_id``.

        Arrival order, not timestamp order, and that distinction is the whole
        point. A probe draining a buffer delivers readings that are *old by
        timestamp* and *new by arrival*. A watermark on ``epoch`` cannot tell
        those apart from history already examined — so the backlog flushed right
        after a hub restart, which is precisely when a probe has a backlog, looks
        like nothing new and is never scanned. ``id`` is AUTOINCREMENT, so it
        answers "what has landed since I last looked" no matter what the readings
        claim about when they were taken.

        ``limit`` bounds the sweep by arrival too, so what it drops is the newest
        (which the live evaluator sees anyway) rather than the oldest, which is
        where a backfilled excursion begins.

        ``site`` follows the same convention as every other reader here: ``''``
        is this hub's own probes, a name is one forwarding store, ``None`` is all
        of them. The alert monitor passes ``''`` — see the note on
        :meth:`latest_per_probe` about why head office must not re-judge another
        hub's probes.
        """
        clauses, params = ["id > ?"], [int(after_id)]
        if site is not None:
            clauses.append("site = ?")
            params.append(site)
        sql = ("SELECT id, probe_id, epoch, temperature_c FROM readings "
               f"WHERE {' AND '.join(clauses)} ORDER BY id ASC")
        if limit is not None:
            sql += " LIMIT ?"
            params.append(int(limit))
        rows = self._conn().execute(sql, params).fetchall()
        return [(r["id"], r["probe_id"], r["epoch"], r["temperature_c"]) for r in rows]

    def list_events(self, limit: int = 50, window_seconds: Optional[int] = None,
                    kinds: Optional[Iterable[str]] = None,
                    exclude_kinds: Optional[Iterable[str]] = None,
                    site: Optional[str] = None) -> list:
        """Most recent alert-lifecycle events as dict rows, newest first.

        Row keys: ``timestamp, epoch, kind, probe_id, temperature_c, limit_c,
        site``. ``site`` filters to the hub that recorded them — ``''`` for this
        hub's own, a store label for the events that store forwarded, ``None``
        (default) for every hub at once.
        ``window_seconds`` bounds the log to a rolling window (index-backed via
        ``idx_events_epoch``); ``limit`` caps the payload for the UI/API.
        ``kinds`` restricts the result to those event kinds (whitelist);
        ``exclude_kinds`` drops them (blacklist). Filtering by kind *in SQL* —
        before the ``LIMIT`` — is what lets a caller pull, say, the newest N
        *alerts* without a flapping probe's online/offline churn evicting them
        from the fetch window.
        """
        conn = self._conn()
        cutoff = self._cutoff(window_seconds)
        clauses: list = []
        params: list = []
        if cutoff is not None:
            clauses.append("epoch >= ?")
            params.append(cutoff)
        kind_list = [str(k) for k in kinds] if kinds else None
        if kind_list:
            clauses.append(f"kind IN ({','.join('?' * len(kind_list))})")
            params.extend(kind_list)
        drop_list = [str(k) for k in exclude_kinds] if exclude_kinds else None
        if drop_list:
            clauses.append(f"kind NOT IN ({','.join('?' * len(drop_list))})")
            params.extend(drop_list)
        if site is not None:
            clauses.append("site = ?")
            params.append(site)
        where = _where(clauses)
        try:
            cap = max(1, int(limit))
        except (TypeError, ValueError):
            cap = 50
        rows = conn.execute(
            f"SELECT ts AS timestamp, epoch, kind, probe_id, temperature_c, limit_c, site "
            f"FROM events {where} ORDER BY epoch DESC, id DESC LIMIT ?",
            tuple(params) + (cap,),
        ).fetchall()
        return [dict(r) for r in rows]

    # -- maintenance -----------------------------------------------------------
    def purge_older_than(self, days: int) -> int:
        """Delete readings older than ``days`` days. Returns rows removed.

        Guarded against a wall clock that has jumped FORWARD, the mirror of the
        risk :meth:`delete_future_readings` guards below. This deletes on
        ``epoch < now - days``, so a clock reading far in the past is harmless
        (the cutoff goes negative and nothing matches) but a clock reading far in
        the future — a bad NTP answer, a VM resumed from a stale snapshot, a
        mistyped date — puts the cutoff beyond every real reading and silently
        deletes the entire history on the next hourly sweep. There is no undo.

        The invariant that catches it without needing a trustworthy clock: a
        retention sweep is meant to trim the tail, so it must never be the thing
        that empties the store. If the cutoff is at or past the newest reading we
        hold, the clock disagrees with the data and we decline rather than guess.

        The one legitimate case this also declines is a hub that has been off
        longer than the retention window, where every row genuinely is expired.
        That resolves itself: once the probes report again the newest epoch is
        current, the cutoff falls back behind it, and the stale rows purge on the
        next sweep. Keeping data too long is recoverable; deleting it is not.
        """
        if not days or int(days) <= 0:
            return 0
        cutoff = int(time.time()) - int(days) * 86400
        conn = self._conn()
        with self._write_lock:
            newest = conn.execute("SELECT MAX(epoch) FROM readings").fetchone()[0]
            if newest is not None and cutoff >= float(newest):
                log.warning(
                    "retention purge skipped: cutoff %d is at or past the newest "
                    "reading (%s) — the system clock looks wrong, or the hub has "
                    "been offline longer than the %d-day retention window. "
                    "Nothing deleted.", cutoff, newest, int(days))
                return 0
            cur = conn.execute("DELETE FROM readings WHERE epoch < ?", (cutoff,))
            conn.commit()
            return cur.rowcount

    def delete_future_readings(self, tolerance_sec: int = 120) -> int:
        """Delete readings stamped implausibly far in the future. Returns rows removed.

        A probe with a bad clock can stamp a reading ahead of real time (see the
        ingest-time clamp in core.storage). Any such row already in the store
        would draw the chart past 'now' and become the bogus 'latest' reading —
        masking a live threshold breach. Run once at startup so an existing
        database self-heals; new readings are clamped at ingest so none recur.

        Guarded against an UNTRUSTWORTHY clock: this runs at startup, and hub
        hardware without a battery-backed RTC (e.g. a Raspberry Pi) can boot to
        1970/epoch-0 or a stale date before NTP syncs. Trusting that low clock
        here would mark the entire legitimate history as "future" and delete it
        irreversibly, so when the wall clock is implausibly early the purge is
        skipped (real data is never at risk; the rare legacy-future cleanup just
        waits for the clock to sync).
        """
        now = int(time.time())
        if now < _CLOCK_TRUSTWORTHY_FLOOR_EPOCH:
            return 0
        cutoff = now + int(tolerance_sec)
        conn = self._conn()
        with self._write_lock:
            cur = conn.execute("DELETE FROM readings WHERE epoch > ?", (cutoff,))
            conn.commit()
            return cur.rowcount

    def delete_probe(self, probe_id: str) -> int:
        """Delete every reading for a single probe. Returns rows removed.

        Used by "remove device" so a decommissioned/test probe's history stops
        showing up in the dashboard, stats and CSV export. An empty probe_id is
        a no-op guard so a blank id can't wipe the unlabelled bucket by accident.
        """
        pid = (probe_id or "").strip()
        if not pid:
            return 0
        conn = self._conn()
        with self._write_lock:
            cur = conn.execute("DELETE FROM readings WHERE probe_id = ?", (pid,))
            conn.commit()
            return cur.rowcount

    def backup(self, dest_path: str | Path) -> None:
        """Write a consistent snapshot of the database to ``dest_path``."""
        dest = sqlite3.connect(str(dest_path))
        try:
            with self._write_lock:
                self._conn().backup(dest)
        finally:
            dest.close()

    # -- export ----------------------------------------------------------------
    # Excel refuses to open a worksheet with more than 1,048,576 rows (including
    # the header); the friendly .xlsx export guards against this instead of
    # producing a file Excel silently truncates or rejects.
    XLSX_MAX_ROWS = 1_048_576

    def _export_where(self, window_seconds, probe_id, start_epoch, end_epoch,
                      site=None):
        """Build the shared ``WHERE`` clause + params for every export variant.

        Filters (all AND-ed, any may be omitted): a rolling ``window_seconds``, a
        single ``probe_id``, an absolute ``start_epoch``/``end_epoch`` range, and
        a ``site``.

        ``site`` follows the convention :func:`_reading_filters` uses, NOT the
        truthiness test applied to ``probe_id``: ``None`` means every site,
        ``''`` means this hub's own probes, and a name means one store. Testing
        it for truth would make ``site=''`` silently mean "no filter", which is
        the exact confusion ``test_empty_site_means_this_hubs_own_probes_in_every_query``
        exists to prevent.
        """
        clauses, params_list = [], []
        cutoff = self._cutoff(window_seconds)
        if cutoff is not None:
            clauses.append("epoch >= ?"); params_list.append(cutoff)
        if start_epoch is not None:
            clauses.append("epoch >= ?"); params_list.append(float(start_epoch))
        if end_epoch is not None:
            # float(), not int(): a `to=YYYY-MM-DD` filter resolves to the
            # fractional end-of-day epoch (…23:59:59.999999) so that
            # `epoch <= bound` keeps the final second's millisecond rows. int()
            # truncated the bound to …59 and silently dropped 23:59:59.001–.999.
            clauses.append("epoch <= ?"); params_list.append(float(end_epoch))
        if probe_id:
            clauses.append("probe_id = ?"); params_list.append(probe_id)
        if site is not None:
            clauses.append("site = ?"); params_list.append(site)
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        return where, tuple(params_list)

    _utc_string = staticmethod(utc_iso)

    @staticmethod
    def _local_date_time(ts, epoch):
        """Split a row into ``(date, time)`` objects in the hub's local wall clock.

        Prefers the stored local ISO ``ts`` (what the probe/hub recorded); falls
        back to the epoch in local time if ``ts`` is malformed. Seconds
        resolution — sub-second precision is preserved in the UTC column.
        """
        try:
            dt = datetime.datetime.fromisoformat(str(ts).replace("Z", ""))
        except (ValueError, TypeError):
            dt = datetime.datetime.fromtimestamp(float(epoch))
        return dt.date(), dt.time().replace(microsecond=0)

    def export_csv(self, file_obj, name_map: Optional[dict] = None,
                   window_seconds: Optional[int] = None,
                   probe_id: Optional[str] = None,
                   start_epoch: Optional[int] = None,
                   end_epoch: Optional[int] = None,
                   site: Optional[str] = None) -> int:
        """Write readings to a file-like object as CSV. Returns the row count.

        This is the canonical/system-of-record export: ISO-8601 timestamps and
        every column. Every row carries both the stored local ``timestamp`` and an
        unambiguous ``timestamp_utc`` derived from the row's epoch, so exported
        data stays correct across machines and DST changes.

        ``name_map`` adds the friendly ``probe`` name the spreadsheet export has
        always carried. Nobody reading "Setpoint-000092 was at 4.5 C" knows which
        appliance that is, and the Download CSV button had no way to say — the
        column existed on one export and not the other, for no reason a reader
        could see. Conditional on a name actually being set, exactly like ``site``
        and the humidity pair: a hub that has named nothing writes the file it
        always wrote, byte for byte.
        """
        conn = self._conn()
        names = name_map or {}
        where, params = self._export_where(window_seconds, probe_id, start_epoch, end_epoch, site)
        # A `site` column appears only on a hub that actually holds forwarded
        # readings. Without it, head office's export of six stores is rows of
        # "walkin" at 4 C, -18 C and 21 C with nothing to tell them apart — and
        # this file is the artefact that goes to an inspector. Adding it
        # unconditionally would instead change the shape of every existing
        # customer's export and break whatever imports it, so it is conditional,
        # exactly like the dashboard's site picker.
        multi = bool(self.sites())
        # str() first: config values are user-editable and normalize_config does
        # not coerce probe_names, so a numeric name reached here and .strip()
        # raised -- breaking the canonical export while the Excel ones, which
        # funnel every value through _csv_safe/_xlsx_safe, were fine.
        named = any(str(names.get(k) or "").strip() for k in names)
        writer = _csv.writer(file_obj)
        header = ["timestamp", "timestamp_utc", "temperature_c", "temperature_f",
                  "probe_id"]
        if named:
            header.append("probe")
        header += ["humidity_pct", "vpd_kpa"]
        if multi:
            header.append("site")
        writer.writerow(header)
        n = 0
        for r in conn.execute(
            f"SELECT ts, epoch, temperature_c, temperature_f, probe_id, humidity_pct, "
            f"vpd_kpa, site FROM readings {where} ORDER BY epoch ASC",
            params,
        ):
            hum = "" if r["humidity_pct"] is None else f"{r['humidity_pct']:.2f}"
            vpd = "" if r["vpd_kpa"] is None else f"{r['vpd_kpa']:.3f}"
            row = [_csv_safe(r["ts"]), self._utc_string(r["epoch"]),
                   f"{r['temperature_c']:.3f}", f"{r['temperature_f']:.3f}",
                   _csv_safe(r["probe_id"])]
            if named:
                # Fall back to the id, never blank: a half-filled column reads as
                # missing data rather than as "this one was never named".
                row.append(_csv_safe(names.get(r["probe_id"]) or r["probe_id"]))
            row += [hum, vpd]
            if multi:
                row.append(_csv_safe(r["site"] or "(this hub)"))
            writer.writerow(row)
            n += 1
        return n

    # Column headers shared by the two Excel-friendly variants (CSV + .xlsx).
    _FRIENDLY_HEADERS = ["date", "time", "probe", "temperature_c", "temperature_f",
                         "probe_id", "timestamp_utc"]

    def has_humidity(self, window_seconds: Optional[int] = None,
                     probe_id: Optional[str] = None,
                     start_epoch: Optional[int] = None,
                     end_epoch: Optional[int] = None,
                     site: Optional[str] = None) -> bool:
        """Whether any reading MATCHING THESE FILTERS carries humidity.

        Gates the humidity/VPD columns on the spreadsheet exports the same way
        ``sites()`` gates the site column: present when the data is actually
        there, absent otherwise, so a plain temperature deployment's export is
        byte-identical to what it has always been.

        It takes the export's own filters because the question the export needs
        answered is about the file being written, not about the hub. Asked
        globally, one grow probe anywhere in the store put two permanently-empty
        columns on every export of every other probe — including the single-
        freezer export a restaurant hands to an inspector, which
        ``export_friendly_csv``'s docstring promises has "unused humidity/VPD
        dropped".

        Called with no arguments it is the same single indexed EXISTS as before.
        With filters it is bounded by the rows the export is about to read
        anyway.
        """
        where, params = self._export_where(window_seconds, probe_id,
                                           start_epoch, end_epoch, site)
        clause = f"{where} AND humidity_pct IS NOT NULL" if where \
            else "WHERE humidity_pct IS NOT NULL"
        row = self._conn().execute(
            f"SELECT 1 FROM readings {clause} LIMIT 1", params).fetchone()
        return row is not None

    def export_friendly_csv(self, file_obj, name_map: Optional[dict] = None,
                            window_seconds: Optional[int] = None,
                            probe_id: Optional[str] = None,
                            start_epoch: Optional[int] = None,
                            end_epoch: Optional[int] = None,
                            site: Optional[str] = None) -> int:
        """Write an Excel-friendly CSV. Returns the row count.

        Same filters as :meth:`export_csv`, but reshaped for people who open the
        file directly in a spreadsheet:

        * ``date`` and ``time`` are split into separate columns (local wall
          clock) so Excel/Sheets parse each as a real date/time value and sort,
          filter and pivot natively — an ISO ``...T...``-with-milliseconds string
          is imported as text and won't.
        * ``probe`` shows the friendly name set in the dashboard (``name_map``),
          with the raw ``probe_id`` kept alongside for disambiguation.
        * ``humidity_pct``/``vpd_kpa`` appear only when the hub actually holds
          them. They used to be dropped unconditionally while the docstring said
          "unused", so a grow deployment — the niche those columns exist for —
          lost humidity and VPD from the one export a non-technical user opens.
        * ``timestamp_utc`` is retained (full precision) so the exact,
          machine-independent instant is never lost.
        """
        names = name_map or {}
        conn = self._conn()
        where, params = self._export_where(window_seconds, probe_id, start_epoch, end_epoch, site)
        # See export_csv: `site` only on a hub that holds forwarded readings, and
        # it matters MORE here — this variant leads with the friendly name, and
        # six stores all calling a probe "Walk-in" is the ordinary case.
        multi = bool(self.sites())
        # The export's OWN filters, so a temperature-only probe's export does
        # not carry two empty columns just because a grow probe exists elsewhere
        # in the same store.
        humid = self.has_humidity(window_seconds, probe_id, start_epoch, end_epoch, site)
        writer = _csv.writer(file_obj)
        writer.writerow(self._FRIENDLY_HEADERS
                        + (["humidity_pct", "vpd_kpa"] if humid else [])
                        + (["site"] if multi else []))
        n = 0
        for r in conn.execute(
            f"SELECT ts, epoch, temperature_c, temperature_f, probe_id, site, "
            f"humidity_pct, vpd_kpa FROM readings {where} ORDER BY epoch ASC",
            params,
        ):
            d, t = self._local_date_time(r["ts"], r["epoch"])
            pid = r["probe_id"] or ""
            friendly = names.get(pid, pid) if isinstance(names, dict) else pid
            row = [d.isoformat(), t.isoformat(), _csv_safe(friendly),
                   f"{r['temperature_c']:.3f}", f"{r['temperature_f']:.3f}",
                   _csv_safe(pid), self._utc_string(r["epoch"])]
            if humid:
                row.append("" if r["humidity_pct"] is None else f"{r['humidity_pct']:.2f}")
                row.append("" if r["vpd_kpa"] is None else f"{r['vpd_kpa']:.3f}")
            if multi:
                row.append(_csv_safe(r["site"] or "(this hub)"))
            writer.writerow(row)
            n += 1
        return n

    def count_readings(self, window_seconds: Optional[int] = None,
                       probe_id: Optional[str] = None,
                       start_epoch: Optional[int] = None,
                       end_epoch: Optional[int] = None,
                       site: Optional[str] = None) -> int:
        """Count readings matching the export filters (used for the .xlsx guard)."""
        conn = self._conn()
        where, params = self._export_where(window_seconds, probe_id, start_epoch, end_epoch, site)
        row = conn.execute(f"SELECT COUNT(*) AS n FROM readings {where}", params).fetchone()
        return int(row["n"]) if row else 0

    def export_xlsx(self, file_obj, name_map: Optional[dict] = None,
                    window_seconds: Optional[int] = None,
                    probe_id: Optional[str] = None,
                    start_epoch: Optional[int] = None,
                    end_epoch: Optional[int] = None,
                    site: Optional[str] = None) -> int:
        """Write a native .xlsx workbook. Returns the row count.

        Same reshaped columns as :meth:`export_friendly_csv`, but as a real Excel
        file so ``date``/``time`` are true date/time cells and the temperatures
        are real numbers — the user double-clicks and everything is already
        typed, sorted, filterable (auto-filter) with a frozen header row.

        Streams via openpyxl's ``write_only`` mode so a long log doesn't buffer
        in memory. Raises :class:`ExportTooLargeForXlsx` when the result would
        exceed Excel's row limit (the caller should offer CSV instead), and
        ``ImportError`` if openpyxl isn't installed.
        """
        from openpyxl import Workbook  # lazy: optional dependency
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        names = name_map or {}
        total = self.count_readings(window_seconds, probe_id, start_epoch, end_epoch, site)
        if total > self.XLSX_MAX_ROWS - 1:  # -1 leaves room for the header row
            raise ExportTooLargeForXlsx(total, self.XLSX_MAX_ROWS - 1)

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Readings")
        ws.freeze_panes = "A2"  # keep the header visible while scrolling
        # See export_csv for why `site` is conditional. The auto-filter and
        # column widths have to track it, or the dropdowns stop one column short
        # of the data and the new column renders at default width.
        multi = bool(self.sites())
        # Same gate as export_friendly_csv, filters and all.
        humid = self.has_humidity(window_seconds, probe_id, start_epoch, end_epoch, site)
        headers = (self._FRIENDLY_HEADERS
                   + (["humidity_pct", "vpd_kpa"] if humid else [])
                   + (["site"] if multi else []))
        widths = ([12, 11, 22, 15, 15, 20, 26]
                  + ([14, 12] if humid else [])
                  + ([16] if multi else []))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Filter dropdowns over the whole populated table (header + data rows).
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{total + 1}"

        bold = Font(bold=True)
        header = []
        for label in headers:
            c = WriteOnlyCell(ws, value=label)
            c.font = bold
            header.append(c)
        ws.append(header)

        conn = self._conn()
        where, params = self._export_where(window_seconds, probe_id, start_epoch, end_epoch, site)
        n = 0
        for r in conn.execute(
            f"SELECT ts, epoch, temperature_c, temperature_f, probe_id, site, "
            f"humidity_pct, vpd_kpa FROM readings {where} ORDER BY epoch ASC",
            params,
        ):
            d, t = self._local_date_time(r["ts"], r["epoch"])
            pid = r["probe_id"] or ""
            friendly = names.get(pid, pid) if isinstance(names, dict) else pid
            date_cell = WriteOnlyCell(ws, value=d)
            date_cell.number_format = "yyyy-mm-dd"
            time_cell = WriteOnlyCell(ws, value=t)
            time_cell.number_format = "hh:mm:ss"
            c_cell = WriteOnlyCell(ws, value=round(float(r["temperature_c"]), 3))
            c_cell.number_format = "0.000"
            f_cell = WriteOnlyCell(ws, value=round(float(r["temperature_f"]), 3))
            f_cell.number_format = "0.000"
            row = [date_cell, time_cell, _xlsx_safe(friendly),
                   c_cell, f_cell, _xlsx_safe(pid), self._utc_string(r["epoch"])]
            if humid:
                # Numbers, not text: a grower filtering or charting VPD in Excel
                # needs real numeric cells.
                for val, fmt in ((r["humidity_pct"], "0.00"), (r["vpd_kpa"], "0.000")):
                    cell = WriteOnlyCell(ws, value=None if val is None else float(val))
                    cell.number_format = fmt
                    row.append(cell)
            if multi:
                row.append(_xlsx_safe(r["site"] or "(this hub)"))
            ws.append(row)
            n += 1
        wb.save(file_obj)
        return n



    # ---- multi-site forwarding support (see core/forwarder.py) ---------------

    def meta_get(self, key: str, default: str = "") -> str:
        """Read a durable scratchpad value. Missing key -> ``default``."""
        row = self._conn().execute("SELECT v FROM meta WHERE k = ?", (key,)).fetchone()
        return row[0] if row else default

    def meta_set(self, key: str, value: str) -> None:
        conn = self._conn()
        with self._write_lock:
            conn.execute("INSERT INTO meta (k, v) VALUES (?, ?) "
                         "ON CONFLICT(k) DO UPDATE SET v = excluded.v", (key, str(value)))
            conn.commit()

    def rows_after(self, after_id: int, limit: int = 500, local_only: bool = True):
        """Readings with ``id > after_id``, oldest first, for upstream forwarding.

        ``local_only`` skips rows that arrived here *from* another hub (site set),
        which is what stops two hubs pointed at each other from ping-ponging the
        same readings forever. A store hub only ever forwards its own.

        Ordering by the autoincrement id — not epoch — is deliberate: it is the
        insertion order, so a probe draining a backlog of old readings still gets
        forwarded, where an epoch-based cursor would skip anything older than the
        high-water mark.
        """
        sql = ("SELECT id, ts, epoch, temperature_c, temperature_f, probe_id, "
               "humidity_pct, vpd_kpa, battery_pct FROM readings WHERE id > ?")
        if local_only:
            sql += " AND (site IS NULL OR site = '')"
        sql += " ORDER BY id LIMIT ?"
        return self._conn().execute(sql, (int(after_id), int(limit))).fetchall()

    def events_after(self, after_id: int, limit: int = 500, local_only: bool = True):
        """Events with ``id > after_id``, oldest first, for upstream forwarding.

        Same contract as :meth:`rows_after` — insertion-order cursor, and
        ``local_only`` excludes events that arrived here from another hub so two
        hubs pointed at each other cannot ping-pong them.
        """
        sql = ("SELECT id, ts, epoch, kind, probe_id, temperature_c, limit_c "
               "FROM events WHERE id > ?")
        if local_only:
            sql += " AND (site IS NULL OR site = '')"
        sql += " ORDER BY id LIMIT ?"
        return self._conn().execute(sql, (int(after_id), int(limit))).fetchall()

    def count_local_after(self, after_id: int) -> int:
        """How many locally-ingested readings are past the forwarding cursor.

        The Settings page shows this as the backlog waiting to reach head office,
        so "it says enabled but is anything happening?" has an answer.
        """
        row = self._conn().execute(
            "SELECT COUNT(*) FROM readings WHERE id > ? AND (site IS NULL OR site = '')",
            (int(after_id),)).fetchone()
        return int(row[0]) if row else 0

    def probe_ids(self):
        """Every distinct probe id in the store, ascending.

        Loose index scan for the same reason :meth:`sites` uses one: SELECT
        DISTINCT over a low-cardinality column visits every row, and this is on
        the dashboard's 5 s path. ``idx_readings_probe_epoch_site_uniq`` leads
        with ``probe_id``, so MIN-then-MIN-of-what-is-greater walks it a probe at
        a time.
        """
        rows = self._conn().execute(
            "WITH RECURSIVE t(p) AS ("
            "  SELECT MIN(probe_id) FROM readings"
            "  UNION ALL"
            "  SELECT (SELECT MIN(probe_id) FROM readings WHERE probe_id > t.p)"
            "    FROM t WHERE t.p IS NOT NULL"
            ") SELECT p FROM t WHERE p IS NOT NULL ORDER BY p").fetchall()
        return [r[0] for r in rows]

    def sites(self):
        """Distinct non-empty site names present in the store, alphabetically.

        A "loose index scan", not ``SELECT DISTINCT``. The dashboard's site picker
        re-reads this on every 5 s refresh tick, and plain DISTINCT over a
        low-cardinality column is a full table scan — 883 ms on a six-store HQ
        with 1.5M rows, every tick, on the callback thread. The recursive form
        walks ``idx_readings_site`` one key at a time (MIN, then MIN of what is
        greater), so the cost is O(sites x log N) instead of O(rows): the same
        query on the same data measured 0.31 ms.

        ``site != ''`` in both arms is not just a filter — it is what lets the
        PARTIAL index serve the query, since a partial index is only usable when
        the query repeats its WHERE clause.
        """
        rows = self._conn().execute(
            "WITH RECURSIVE t(s) AS ("
            "  SELECT MIN(site) FROM readings WHERE site != ''"
            "  UNION ALL"
            "  SELECT (SELECT MIN(site) FROM readings WHERE site > t.s AND site != '')"
            "    FROM t WHERE t.s IS NOT NULL"
            ") SELECT s FROM t WHERE s IS NOT NULL ORDER BY s").fetchall()
        return [r[0] for r in rows]


def migrate_csv_if_present(db: "Database", csv_path: str | Path) -> int:
    """One-time import of a legacy ``temperature_log.csv`` into the database.

    Runs only when the database is empty, so it is safe to call on every start.
    Returns the number of rows imported (0 if nothing to do).
    """
    csv_path = Path(csv_path)
    if not csv_path.exists() or db.count() > 0:
        return 0
    rows = []
    try:
        with csv_path.open("r", encoding="utf-8", newline="") as f:
            reader = _csv.DictReader(f)
            for row in reader:
                ts = (row.get("timestamp") or "").strip()
                if not ts:
                    continue
                try:
                    t_c = float(row.get("temperature_c"))
                except (TypeError, ValueError):
                    continue
                try:
                    t_f = float(row.get("temperature_f"))
                except (TypeError, ValueError):
                    t_f = c_to_f(t_c)
                rows.append((ts, t_c, t_f, (row.get("probe_id") or "").strip()))
        return db.bulk_insert(rows)
    except Exception:
        # Partial import is still useful; insert whatever parsed cleanly.
        return db.bulk_insert(rows)
