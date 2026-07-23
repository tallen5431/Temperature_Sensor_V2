"""Tests for the SQLite reading store (core.db)."""
import datetime
import io
import time

import pytest

from core.db import Database, iso_to_epoch, migrate_csv_if_present


def _iso(dt: datetime.datetime) -> str:
    return dt.replace(microsecond=0).isoformat()


@pytest.fixture()
def db(tmp_path):
    return Database(tmp_path / "test.db")


def test_empty_database(db):
    assert db.count() == 0
    assert db.latest() is None
    stats = db.window_stats()
    assert stats["count"] == 0 and stats["min"] is None
    assert db.window_df().empty


def test_append_and_latest(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(seconds=10)), 20.0, 68.0, "probeA")
    db.append(_iso(now), 22.5, 72.5, "probeA")
    assert db.count() == 2
    latest = db.latest()
    assert latest["temperature_c"] == 22.5
    assert latest["probe_id"] == "probeA"


def test_window_filtering(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(hours=48)), 10.0, 50.0, "p")  # old
    db.append(_iso(now - datetime.timedelta(minutes=30)), 20.0, 68.0, "p")  # recent
    db.append(_iso(now), 21.0, 69.8, "p")  # recent

    all_rows = db.window_df()
    assert len(all_rows) == 3

    last_hour = db.window_df(window_seconds=3600)
    assert len(last_hour) == 2  # the 48h-old row is excluded


def test_window_stats_accurate(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(minutes=3)), 10.0, 50.0, "p")
    db.append(_iso(now - datetime.timedelta(minutes=2)), 30.0, 86.0, "p")
    db.append(_iso(now - datetime.timedelta(minutes=1)), 20.0, 68.0, "p")
    stats = db.window_stats(window_seconds=3600)
    assert stats["count"] == 3
    assert stats["min"] == 10.0
    assert stats["max"] == 30.0
    assert stats["avg"] == pytest.approx(20.0)
    # min/max timestamps point at the correct rows
    assert stats["min_ts"].endswith(_iso(now - datetime.timedelta(minutes=3))[-8:])


def test_downsampling_caps_points_but_not_stats(db):
    now = datetime.datetime.now()
    base = now - datetime.timedelta(hours=1)
    for i in range(1000):
        db.append(_iso(base + datetime.timedelta(seconds=i)), float(i % 50), 0.0, "p")
    df = db.window_df(window_seconds=7200, max_points=100)
    assert len(df) <= 100  # plot is downsampled
    stats = db.window_stats(window_seconds=7200)
    assert stats["count"] == 1000  # stats use the full set
    assert stats["max"] == 49.0


def test_downsampling_keeps_every_probe_with_noncontiguous_ids(db):
    # Regression: the old `id % stride` sampler could return a blank chart or drop
    # whole probes when primary keys weren't contiguous (interleaved writers, or a
    # deleted probe leaving gaps). Interleave two probes so ids alternate, delete
    # one so only odd ids survive, then downsample hard.
    now = datetime.datetime.now()
    base = now - datetime.timedelta(hours=1)
    for i in range(600):
        db.append(_iso(base + datetime.timedelta(seconds=i)), float(i % 40), 0.0, "A")
        db.append(_iso(base + datetime.timedelta(seconds=i)), float(i % 40), 0.0, "B")
    db.delete_probe("B")  # survivors ("A") now all have even OR odd ids only
    df = db.window_df(window_seconds=7200, max_points=50)
    assert not df.empty                         # never a blank chart
    assert set(df["probe_id"]) == {"A"}         # the surviving probe is present
    assert len(df) <= 50

    # Three probes reporting round-robin (ids in fixed residue classes mod 3):
    # all three must still appear after downsampling.
    d3 = Database(str(db.path) + ".three")
    for i in range(300):
        for pid in ("X", "Y", "Z"):
            d3.append(_iso(base + datetime.timedelta(seconds=i)), float(i % 30), 0.0, pid)
    df3 = d3.window_df(window_seconds=7200, max_points=30)
    assert set(df3["probe_id"]) == {"X", "Y", "Z"}


def test_latest_per_probe(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(seconds=20)), 19.0, 0.0, "A")
    db.append(_iso(now - datetime.timedelta(seconds=10)), 25.0, 0.0, "A")
    db.append(_iso(now - datetime.timedelta(seconds=5)), 30.0, 0.0, "B")
    latest = db.latest_per_probe()
    by_probe = {r["probe_id"]: r["temperature_c"] for _, r in latest.iterrows()}
    assert by_probe == {"A": 25.0, "B": 30.0}


def test_latest_per_probe_window_and_same_second_tie(db):
    now = datetime.datetime.now()
    ts = _iso(now)
    db.append(ts, 20.0, 0.0, "A")
    db.append(ts, 21.0, 0.0, "A")  # same second: the later insert (higher id) wins
    db.append(_iso(now - datetime.timedelta(hours=2)), 5.0, 0.0, "OLD")
    latest = db.latest_per_probe(window_seconds=3600)
    by_probe = {r["probe_id"]: r["temperature_c"] for _, r in latest.iterrows()}
    assert by_probe == {"A": 21.0}  # OLD is outside the window entirely


def test_battery_stored_and_exposed(db):
    now = datetime.datetime.now()
    db.append(_iso(now), 20.0, 68.0, "A", battery=87.5)
    db.append(_iso(now), 4.0, 39.2, "B")  # no battery telemetry -> NULL
    latest = db.latest_per_probe()
    by_probe = {r["probe_id"]: r["battery_pct"] for _, r in latest.iterrows()}
    assert by_probe["A"] == 87.5
    # SQL NULL surfaces as None/NaN in the frame (matches humidity_pct; the API
    # maps either to JSON null via _num).
    assert by_probe["B"] is None or by_probe["B"] != by_probe["B"]
    rows = {r["probe_id"]: r for r in db.fetch_readings()}
    assert rows["A"]["battery_pct"] == 87.5
    assert rows["B"]["battery_pct"] is None


def test_has_probe_prefix(db):
    assert db.has_probe_prefix("DEMO-") is False  # empty store
    now = datetime.datetime.now()
    db.append(_iso(now), 4.0, 39.2, "DEMO-Fridge")
    db.append(_iso(now), 20.0, 68.0, "REAL-1")
    assert db.has_probe_prefix("DEMO-") is True
    assert db.has_probe_prefix("XYZ-") is False
    # An empty prefix degenerates to "any reading at all".
    assert db.has_probe_prefix("") is True


def test_has_probe_prefix_is_index_backed(db):
    # The whole point of the helper is an O(log N) probe: the GLOB prefix must
    # rewrite to a range seek on idx_readings_probe_epoch, never a table scan.
    # (If a SQLite build stops applying the optimisation, switch the helper to
    # the equivalent  probe_id >= 'DEMO-' AND probe_id < 'DEMO.'  range form.)
    plan = " ".join(
        str(r[3]) for r in db._conn().execute(
            "EXPLAIN QUERY PLAN SELECT EXISTS"
            "(SELECT 1 FROM readings WHERE probe_id GLOB ?)", ("DEMO-*",))
    )
    assert "idx_readings_probe_epoch" in plan


def test_record_and_list_events(db):
    now = datetime.datetime.now()
    db.record_event("high", "A", temperature_c=30.0, limit=25.0,
                    ts=_iso(now - datetime.timedelta(seconds=20)))
    db.record_event("recovery", "A", temperature_c=24.0, limit=25.0,
                    ts=_iso(now - datetime.timedelta(seconds=10)))
    db.record_event("offline", "B", ts=_iso(now))
    events = db.list_events()
    assert [e["kind"] for e in events] == ["offline", "recovery", "high"]  # newest first
    assert set(events[0]) == {"timestamp", "epoch", "kind", "probe_id",
                              "temperature_c", "limit_c"}
    assert events[2]["probe_id"] == "A"
    assert events[2]["temperature_c"] == 30.0 and events[2]["limit_c"] == 25.0
    assert events[0]["temperature_c"] is None and events[0]["limit_c"] is None
    # limit caps the payload, keeping the newest rows
    assert [e["kind"] for e in db.list_events(limit=2)] == ["offline", "recovery"]


def test_list_events_window(db):
    now = datetime.datetime.now()
    db.record_event("low", "A", ts=_iso(now - datetime.timedelta(hours=2)))
    db.record_event("online", "A", ts=_iso(now))
    assert len(db.list_events()) == 2
    recent = db.list_events(window_seconds=3600)
    assert [e["kind"] for e in recent] == ["online"]


def test_record_event_tolerates_bad_input(db):
    # Unparseable numerics are coerced to NULL, never raised to the caller.
    db.record_event("high", "A", temperature_c="not-a-number", limit=object())
    ev = db.list_events()[0]
    assert ev["kind"] == "high" and ev["temperature_c"] is None and ev["limit_c"] is None
    # A blank kind is skipped; a garbage ts falls back to ~now via iso_to_epoch.
    db.record_event("", "A")
    assert len(db.list_events()) == 1
    db.record_event("rate", None, ts="not-a-date")
    ev = db.list_events()[0]
    assert ev["kind"] == "rate" and ev["probe_id"] == ""
    assert abs(ev["epoch"] - time.time()) < 5


def test_record_event_defaults_ts_to_now(db):
    db.record_event("offline", "A")
    ev = db.list_events()[0]
    assert abs(ev["epoch"] - time.time()) < 5
    # the stored timestamp is a local ISO-seconds string
    datetime.datetime.fromisoformat(ev["timestamp"])


def test_export_csv_roundtrip(db):
    now = datetime.datetime.now()
    db.append(_iso(now), 22.123, 71.821, "probeX")
    buf = io.StringIO()
    n = db.export_csv(buf)
    assert n == 1
    content = buf.getvalue()
    assert "timestamp,timestamp_utc,temperature_c,temperature_f,probe_id" in content
    assert "probeX" in content
    assert "22.123" in content
    # every data row carries an unambiguous UTC timestamp ending in Z
    assert content.splitlines()[1].split(",")[1].endswith("Z")


def test_export_csv_filters(db):
    now = datetime.datetime.now()
    old = now - datetime.timedelta(days=3)
    db.append(_iso(old), 10.0, 50.0, "A")     # 3 days ago
    db.append(_iso(now), 20.0, 68.0, "A")     # today
    db.append(_iso(now), 4.0, 39.2, "B")      # today, other probe

    def rows(**kw):
        buf = io.StringIO()
        n = db.export_csv(buf, **kw)
        return n

    assert rows() == 3
    assert rows(probe_id="A") == 2
    assert rows(probe_id="B") == 1
    # absolute date range: only today's readings
    start = int(now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())
    assert rows(start_epoch=start) == 2
    # combine probe + range
    assert rows(probe_id="A", start_epoch=start) == 1
    # end_epoch excludes today
    assert rows(end_epoch=start - 1) == 1  # only the 3-days-ago row


def test_export_friendly_csv_shape(db):
    db.append("2026-07-21T22:45:36.267", -18.5, -1.3, "Setpoint-000079", humidity=55.0)
    buf = io.StringIO()
    n = db.export_friendly_csv(buf, name_map={"Setpoint-000079": "Chest Freezer"})
    assert n == 1
    lines = buf.getvalue().splitlines()
    assert lines[0] == "date,time,probe,temperature_c,temperature_f,probe_id,timestamp_utc"
    cols = lines[1].split(",")
    # date and time are split into separate, spreadsheet-parseable columns
    assert cols[0] == "2026-07-21"
    assert cols[1] == "22:45:36"
    # friendly name is used, raw id kept alongside, humidity/vpd dropped entirely
    assert cols[2] == "Chest Freezer"
    assert cols[5] == "Setpoint-000079"
    assert len(cols) == 7
    assert "humidity" not in lines[0] and "vpd" not in lines[0]
    # full-precision UTC instant is retained
    assert cols[6] == "2026-07-21T22:45:36.267Z"


def test_export_friendly_csv_falls_back_to_raw_id_when_unnamed(db):
    db.append("2026-07-21T09:00:00", 4.0, 39.2, "probeZ")
    buf = io.StringIO()
    db.export_friendly_csv(buf, name_map={})  # no friendly name configured
    assert buf.getvalue().splitlines()[1].split(",")[2] == "probeZ"


def test_export_friendly_csv_formula_injection_guard(db):
    # A malicious friendly name / id starting with a formula char is neutralised.
    db.append("2026-07-21T09:00:00", 4.0, 39.2, "probeZ")
    buf = io.StringIO()
    db.export_friendly_csv(buf, name_map={"probeZ": "=cmd|calc"})
    cell = buf.getvalue().splitlines()[1].split(",")[2]
    assert cell.startswith("'=")  # single-quote guard makes Excel treat it as text


def test_export_friendly_csv_filters(db):
    now = datetime.datetime.now()
    old = now - datetime.timedelta(days=3)
    db.append(_iso(old), 10.0, 50.0, "A")
    db.append(_iso(now), 20.0, 68.0, "A")
    db.append(_iso(now), 4.0, 39.2, "B")

    def rows(**kw):
        buf = io.StringIO()
        return db.export_friendly_csv(buf, **kw)

    assert rows() == 3
    assert rows(probe_id="A") == 2
    start = int(now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())
    assert rows(start_epoch=start) == 2
    assert rows(probe_id="A", start_epoch=start) == 1


def test_count_readings_matches_filters(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(days=3)), 10.0, 50.0, "A")
    db.append(_iso(now), 20.0, 68.0, "A")
    db.append(_iso(now), 4.0, 39.2, "B")
    assert db.count_readings() == 3
    assert db.count_readings(probe_id="A") == 2
    start = int(now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())
    assert db.count_readings(start_epoch=start) == 2


def test_export_xlsx_typed_cells(db):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    db.append("2026-07-21T22:45:36.500", -18.5, -1.3, "Setpoint-000079")
    dest = str(db.path) + ".xlsx"
    with open(dest, "wb") as f:
        n = db.export_xlsx(f, name_map={"Setpoint-000079": "Chest Freezer"})
    assert n == 1

    ws = load_workbook(dest).active
    # header is present, frozen, and an auto-filter spans the populated table
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header == ["date", "time", "probe", "temperature_c",
                      "temperature_f", "probe_id", "timestamp_utc"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:G2"

    row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    date_c, time_c, probe_c, tc, tf, pid, utc = row
    # date/time are real date/time cells, temperatures are real numbers
    assert date_c.data_type == "d" and time_c.data_type == "d"
    assert tc.data_type == "n" and tc.value == -18.5
    assert tf.data_type == "n"
    assert probe_c.value == "Chest Freezer"
    assert pid.value == "Setpoint-000079"
    assert utc.value == "2026-07-21T22:45:36.500Z"


def test_export_xlsx_injection_guard(db):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    db.append("2026-07-21T09:00:00", 4.0, 39.2, "probeZ")
    dest = str(db.path) + ".guard.xlsx"
    with open(dest, "wb") as f:
        db.export_xlsx(f, name_map={"probeZ": "=HYPERLINK(1)"})
    ws = load_workbook(dest).active
    probe_cell = list(ws.iter_rows(min_row=2, max_row=2))[0][2]
    # stored as literal text, never as a formula
    assert probe_cell.data_type == "s"
    assert str(probe_cell.value).startswith("'=")


def test_export_xlsx_row_limit_guard(db):
    from core.db import ExportTooLargeForXlsx

    now = datetime.datetime.now()
    for i in range(5):
        db.append(_iso(now - datetime.timedelta(seconds=i)), float(i), 0.0, "p")
    db.XLSX_MAX_ROWS = 4  # tiny limit: 5 rows > 4 - 1
    with pytest.raises(ExportTooLargeForXlsx) as exc:
        with open(str(db.path) + ".big.xlsx", "wb") as f:
            db.export_xlsx(f)
    assert exc.value.rows == 5


def test_export_csv_canonical_unchanged_by_friendly_addition(db):
    # The system-of-record CSV must keep its exact 7-column ISO shape.
    db.append("2026-07-21T22:45:36.267", -18.5, -1.3, "A", humidity=55.0)
    buf = io.StringIO()
    db.export_csv(buf)
    head = buf.getvalue().splitlines()[0]
    assert head == ("timestamp,timestamp_utc,temperature_c,temperature_f,"
                    "probe_id,humidity_pct,vpd_kpa")


def test_iso_to_epoch_roundtrip():
    now = datetime.datetime.now().replace(microsecond=0)
    epoch = iso_to_epoch(now.isoformat())
    assert abs(epoch - now.timestamp()) < 1.5
    # trailing Z tolerated
    assert iso_to_epoch(now.isoformat() + "Z") == epoch
    # garbage falls back to ~now instead of raising
    assert abs(iso_to_epoch("not-a-date") - time.time()) < 2


def test_last_reading_epoch_per_probe(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(seconds=30)), 10.0, 0.0, "A")
    db.append(_iso(now), 11.0, 0.0, "A")              # newer for A
    db.append(_iso(now - datetime.timedelta(seconds=5)), 20.0, 0.0, "B")
    last = db.last_reading_epoch_per_probe()
    assert set(last.keys()) == {"A", "B"}
    assert last["A"] >= last["B"]  # A's newest is more recent than B's


def test_subsecond_epoch_is_fractional():
    # A millisecond timestamp keeps its sub-second part in the epoch (not floored
    # to a whole second), so high-rate readings stay ordered and distinguishable.
    assert iso_to_epoch("2026-07-21T00:42:04.500") % 1 == 0.5
    assert iso_to_epoch("2026-07-21T00:42:04") % 1 == 0.0  # whole second unchanged


def test_subsecond_readings_ordering_and_export(db):
    # Two readings in the same wall-clock second, 0.5 s apart (a 2 Hz cadence).
    db.append("2026-07-21T00:42:04.000", 20.0, 68.0, "A")
    db.append("2026-07-21T00:42:04.500", 20.5, 68.9, "A")
    # "latest" resolves within the second: the .500 reading wins.
    assert float(db.latest_per_probe().iloc[0]["temperature_c"]) == 20.5
    # The export's UTC column carries the millisecond precision through.
    buf = io.StringIO()
    db.export_csv(buf)
    body = buf.getvalue()
    assert ".500Z" in body and "00:42:04.500" in body


def test_stats_per_probe(db):
    now = datetime.datetime.now()
    # Freezer probe A: cold range; room probe B: warm range.
    for t in (-20.0, -18.0, -16.0):
        db.append(_iso(now), t, 0.0, "A")
    for t in (20.0, 22.0, 24.0):
        db.append(_iso(now), t, 0.0, "B")
    stats = db.stats_per_probe()
    assert set(stats.keys()) == {"A", "B"}
    assert stats["A"]["min"] == -20.0 and stats["A"]["max"] == -16.0
    assert stats["A"]["count"] == 3
    assert abs(stats["A"]["avg"] - (-18.0)) < 1e-6
    assert stats["B"]["min"] == 20.0 and stats["B"]["max"] == 24.0
    assert abs(stats["B"]["avg"] - 22.0) < 1e-6


def test_stats_per_probe_empty(db):
    assert db.stats_per_probe() == {}


def test_window_stats_probe_filter(db):
    now = datetime.datetime.now()
    for t in (-20.0, -18.0, -16.0):
        db.append(_iso(now), t, 0.0, "A")
    for t in (20.0, 22.0, 24.0):
        db.append(_iso(now), t, 0.0, "B")
    alls = db.window_stats()
    assert alls["min"] == -20.0 and alls["max"] == 24.0 and alls["count"] == 6
    a = db.window_stats(probe_id="A")
    assert a["min"] == -20.0 and a["max"] == -16.0 and a["count"] == 3
    assert abs(a["avg"] - (-18.0)) < 1e-6
    assert db.window_stats(probe_id="nope")["count"] == 0


def test_delete_probe(db):
    now = datetime.datetime.now()
    for i in range(4):
        db.append(_iso(now), 20.0 + i, 0.0, "keep")
    for i in range(3):
        db.append(_iso(now), 5.0 + i, 0.0, "gone")
    assert db.count() == 7
    removed = db.delete_probe("gone")
    assert removed == 3
    assert db.count() == 4
    assert set(db.stats_per_probe().keys()) == {"keep"}
    # Deleting again removes nothing; a blank id is a guarded no-op.
    assert db.delete_probe("gone") == 0
    assert db.delete_probe("") == 0
    assert db.count() == 4


def test_bulk_insert(db):
    now = datetime.datetime.now()
    rows = [(_iso(now - datetime.timedelta(seconds=i)), float(i), float(i) * 2, "p")
            for i in range(100)]
    n = db.bulk_insert(rows)
    assert n == 100
    assert db.count() == 100
    assert db.bulk_insert([]) == 0  # empty is a no-op


def test_has_any(db):
    assert db.has_any() is False
    db.append(_iso(datetime.datetime.now()), 20.0, 68.0, "p")
    assert db.has_any() is True


def test_delete_future_readings(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(minutes=5)), 22.0, 0.0, "p")   # past
    db.append(_iso(now), 23.0, 0.0, "p")                                   # now
    db.append(_iso(now + datetime.timedelta(minutes=47)), 23.0, 0.0, "p")  # future glitch
    assert db.count() == 3
    removed = db.delete_future_readings()
    assert removed == 1
    assert db.count() == 2
    # The bogus future row is gone, so "latest" is the real current reading.
    assert db.latest()["temperature_c"] == 23.0
    # A tiny skew (well under tolerance) is preserved.
    db.append(_iso(now + datetime.timedelta(seconds=30)), 24.0, 0.0, "p")
    assert db.delete_future_readings() == 0


def test_purge_older_than(db):
    now = datetime.datetime.now()
    db.append(_iso(now - datetime.timedelta(days=40)), 10.0, 50.0, "p")  # old
    db.append(_iso(now - datetime.timedelta(days=5)), 20.0, 68.0, "p")   # recent
    db.append(_iso(now), 21.0, 69.8, "p")                                # recent
    removed = db.purge_older_than(30)
    assert removed == 1
    assert db.count() == 2
    # 0 days = keep forever (no-op)
    assert db.purge_older_than(0) == 0
    assert db.count() == 2


def test_backup_snapshot(db, tmp_path):
    now = datetime.datetime.now()
    for i in range(5):
        db.append(_iso(now - datetime.timedelta(seconds=i)), float(i), 0.0, "p")
    dest = tmp_path / "backup.db"
    db.backup(dest)
    assert dest.exists()
    # The snapshot is a valid, queryable database with the same row count.
    snap = Database(dest)
    assert snap.count() == 5


def test_migrate_csv(tmp_path):
    csv_path = tmp_path / "legacy.csv"
    csv_path.write_text(
        "timestamp,temperature_c,temperature_f,probe_id\n"
        "2026-01-01T10:00:00,20.0,68.0,p1\n"
        "2026-01-01T10:00:05,21.0,69.8,p1\n"
        "bad,row,here,x\n",  # unparseable temp -> skipped
        encoding="utf-8",
    )
    db = Database(tmp_path / "m.db")
    imported = migrate_csv_if_present(db, csv_path)
    assert imported == 2
    assert db.count() == 2
    # second call is a no-op because the db is no longer empty
    assert migrate_csv_if_present(db, csv_path) == 0
